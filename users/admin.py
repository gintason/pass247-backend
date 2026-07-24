from django.contrib import admin
from django.contrib.auth import get_user_model
from django.contrib.auth.admin import UserAdmin as DefaultUserAdmin
from .models import PasswordReset, UserProfile, UserExamProgress, UserActivity
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone

User = get_user_model()


# ============================================================
# ENHANCED EXPORT ACTION - ALL USERS
# ============================================================
@admin.action(description="📥 Export selected users to Excel")
def export_users_to_excel(modeladmin, request, queryset):
    """Export selected users to a well-formatted Excel file"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Users"
    
    # Define styles
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    premium_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Green tint
    not_premium_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # Orange tint
    
    # Define headers
    headers = [
        "ID", "Username", "Email", "First Name", "Last Name", 
        "Is Staff", "Is Superuser", "Date Joined", "Last Login",
        "Is Premium", "Premium Expiry", "Total Practices", 
        "Total Questions", "Average Score", "Total Points",
        "Phone Number", "Location", "Preferred Exam", 
        "Interest Area", "Profile Bio"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_num, user in enumerate(queryset, 2):
        try:
            profile = user.profile
        except UserProfile.DoesNotExist:
            profile = None
        
        row_data = [
            user.id,
            user.username,
            user.email,
            user.first_name or 'N/A',
            user.last_name or 'N/A',
            'Yes' if user.is_staff else 'No',
            'Yes' if user.is_superuser else 'No',
            user.date_joined.strftime("%Y-%m-%d %H:%M") if user.date_joined else 'N/A',
            user.last_login.strftime("%Y-%m-%d %H:%M") if user.last_login else 'Never',
            'Yes' if (profile and profile.is_premium) else 'No',
            profile.premium_expiry.strftime("%Y-%m-%d") if (profile and profile.premium_expiry) else 'N/A',
            profile.total_practices if profile else 0,
            profile.total_questions_answered if profile else 0,
            f"{profile.average_score:.1f}%" if profile else '0%',
            profile.total_points if profile else 0,
            profile.phone_number if profile else 'N/A',
            profile.location if profile else 'N/A',
            profile.get_preferred_exam_type_display() if profile and profile.preferred_exam_type else 'N/A',
            profile.get_interest_area_display() if profile and profile.interest_area else 'N/A',
            profile.bio[:100] if profile and profile.bio else 'N/A',
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Color code premium users
            if col_num == 10:  # Is Premium column
                if value == 'Yes':
                    cell.fill = premium_fill
                else:
                    cell.fill = not_premium_fill
    
    # Adjust column widths
    column_widths = {
        1: 8,   # ID
        2: 20,  # Username
        3: 30,  # Email
        4: 15,  # First Name
        5: 15,  # Last Name
        6: 10,  # Is Staff
        7: 12,  # Is Superuser
        8: 18,  # Date Joined
        9: 18,  # Last Login
        10: 12, # Is Premium
        11: 15, # Premium Expiry
        12: 15, # Total Practices
        13: 16, # Total Questions
        14: 15, # Average Score
        15: 13, # Total Points
        16: 15, # Phone
        17: 15, # Location
        18: 20, # Preferred Exam
        19: 15, # Interest Area
        20: 30, # Bio
    }
    
    for col_num, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 30
    
    # Freeze the header row
    worksheet.freeze_panes = 'A2'
    
    # Add auto-filter
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(queryset) + 1}"
    
    # Add summary sheet
    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Export Summary"])
    summary_sheet.append(["Generated by:", request.user.username])
    summary_sheet.append(["Date:", timezone.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_sheet.append(["Total Users:", len(queryset)])
    summary_sheet.append(["Premium Users:", sum(1 for u in queryset if hasattr(u, 'profile') and u.profile.is_premium)])
    summary_sheet.append(["Active Today:", sum(1 for u in queryset if u.last_login and u.last_login.date() == timezone.now().date())])
    
    # Style summary sheet
    summary_sheet['A1'].font = Font(bold=True, size=14)
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 40
    
    # Generate response
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f"users_export_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


# ============================================================
# EXPORT ALL USERS ACTION (no selection needed)
# ============================================================
@admin.action(description="📥 Export ALL users to Excel")
def export_all_users_to_excel(modeladmin, request, queryset):
    """Export ALL users in the system to Excel"""
    all_users = User.objects.all().select_related('profile')
    return export_users_to_excel(modeladmin, request, all_users)


# ============================================================
# EXPORT PREMIUM USERS ONLY
# ============================================================
@admin.action(description="⭐ Export premium users to Excel")
def export_premium_users_to_excel(modeladmin, request, queryset):
    """Export only premium users to Excel"""
    premium_users = User.objects.filter(profile__is_premium=True).select_related('profile')
    return export_users_to_excel(modeladmin, request, premium_users)


# ============================================================
# EXPORT USER ACTIVITIES
# ============================================================
@admin.action(description="📊 Export selected users' activities to Excel")
def export_user_activities_to_excel(modeladmin, request, queryset):
    """Export activities for selected users"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "User Activities"
    
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["User", "Activity Type", "Description", "IP Address", "Date"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_num = 2
    for user in queryset:
        activities = UserActivity.objects.filter(user=user).order_by('-created_at')[:100]
        for activity in activities:
            row_data = [
                user.username,
                activity.get_activity_type_display(),
                activity.description,
                activity.ip_address or 'N/A',
                activity.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
            row_num += 1
    
    # Adjust column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 18
    worksheet.column_dimensions['C'].width = 40
    worksheet.column_dimensions['D'].width = 18
    worksheet.column_dimensions['E'].width = 22
    
    worksheet.freeze_panes = 'A2'
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=user_activities_export.xlsx'
    workbook.save(response)
    return response


# ============================================================
# ADMIN MODEL REGISTRATIONS
# ============================================================
@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    list_display = [
        'user', 'phone_number', 'is_premium', 'premium_expiry',
        'total_practices', 'total_questions_answered', 'average_score',
        'total_points', 'interest_area'
    ]
    list_filter = ['is_premium', 'interest_area', 'preferred_exam_type', 'created_at']
    search_fields = ['user__username', 'user__email', 'phone_number', 'location']
    readonly_fields = ['total_practices', 'total_questions_answered', 'average_score', 'total_points']
    
    fieldsets = (
        ('User Info', {
            'fields': ('user', 'phone_number', 'date_of_birth', 'profile_picture')
        }),
        ('Subscription', {
            'fields': ('is_premium', 'premium_expiry')
        }),
        ('Preferences', {
            'fields': ('preferred_exam_type', 'interest_area')
        }),
        ('Statistics', {
            'fields': ('total_practices', 'total_questions_answered', 'average_score', 'total_points')
        }),
        ('Profile', {
            'fields': ('bio', 'location')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )


@admin.register(UserExamProgress)
class UserExamProgressAdmin(admin.ModelAdmin):
    list_display = ['user', 'exam_type', 'subject', 'total_questions_attempted', 'total_correct', 'best_score', 'last_practiced']
    list_filter = ['exam_type', 'subject']
    search_fields = ['user__username', 'subject']


@admin.register(UserActivity)
class UserActivityAdmin(admin.ModelAdmin):
    list_display = ['user', 'activity_type', 'description', 'ip_address', 'created_at']
    list_filter = ['activity_type', 'created_at']
    search_fields = ['user__username', 'description']
    readonly_fields = ['created_at']
    date_hierarchy = 'created_at'


# Unregister the default User admin first
try:
    admin.site.unregister(User)
except admin.sites.NotRegistered:
    pass


@admin.register(User)
class CustomUserAdmin(DefaultUserAdmin):
    list_display = (
        "id", "username", "email", "first_name", "last_name", 
        "is_staff", "date_joined", "last_login", "get_is_premium",
        "get_total_practices"
    )
    search_fields = ("username", "email", "first_name", "last_name")
    list_filter = ("is_staff", "is_superuser", "is_active", "date_joined")
    
    # Add all export actions
    actions = [
        export_users_to_excel,
        export_all_users_to_excel,
        export_premium_users_to_excel,
        export_user_activities_to_excel,
    ]
    
    def get_is_premium(self, obj):
        try:
            return obj.profile.is_premium
        except UserProfile.DoesNotExist:
            return False
    get_is_premium.boolean = True
    get_is_premium.short_description = 'Premium'
    
    def get_total_practices(self, obj):
        try:
            return obj.profile.total_practices
        except UserProfile.DoesNotExist:
            return 0
    get_total_practices.short_description = 'Practices'
    
    # Add fieldsets for better organization
    fieldsets = (
        ('Account Info', {
            'fields': ('username', 'password')
        }),
        ('Personal Info', {
            'fields': ('first_name', 'last_name', 'email')
        }),
        ('Permissions', {
            'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions'),
        }),
        ('Important dates', {
            'fields': ('last_login', 'date_joined')
        }),
    )


admin.site.register(PasswordReset)